import openpyxl as xl

wb = xl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
county_data = {}

us_pop = 0

for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    popul = sheet['D' + str(row)].value

    us_pop += popul

    # агрантия существования ключа для штата. У каждого штата будет свой словарь, в котором будут храниться значения округов и популяции
    county_data.setdefault(state, {})
    county_data[state].setdefault(county, {'pop': 0, 'tracts': 0})

    # каждая строка - один переписной район, поэтому увеличиваем счетчик районов на 1
    county_data[state][county]['tracts'] += 1
    # увеличим численность населения округа на численность населения района
    county_data[state][county]['pop'] += popul

print(county_data)

new_wb = xl.Workbook()
sheet = new_wb.active

sheet['A1'] = 'County'
sheet['B1'] = 'Population'
sheet['C1'] = 'Tracts'

row = 2
for state in county_data:
    for county in county_data[state]:
        sheet.cell(row=row, column=1).value = county
        sheet.cell(row=row, column=2).value = county_data[state][county]['pop']
        sheet.cell(row=row, column=3).value = county_data[state][county]['tracts']
        row += 1

max_row = sheet.max_row
sheet.cell(row=max_row + 1, column=1).value = 'Total population: '
sheet.cell(row=max_row + 1, column=2).value = us_pop

new_wb.save('pop_text.xlsx')
