import openpyxl as xl

wb = xl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {
    'Lemon': 3.07,
    'Daikon': 2.19,
    'Lettuce': 3.45,
}

for rowNumber in range(2, sheet.max_row):
    product = sheet.cell(row=rowNumber, column=1).value
    if product in PRICE_UPDATES:
        sheet.cell(row=rowNumber, column=2).value = PRICE_UPDATES[product]

wb.save('updatedProduceSales.xlsx')
