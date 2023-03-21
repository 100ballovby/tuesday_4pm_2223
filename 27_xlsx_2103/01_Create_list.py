import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='Первый лист')
wb.create_sheet(index=2, title='Средний лист')

sheet = wb['Sheet']

print(wb.sheetnames)

