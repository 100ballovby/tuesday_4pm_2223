import openpyxl as xl


wb = xl.load_workbook('example.xlsx')
print(wb.sheetnames)  # выгрузить названия всех листов таблицы

sheet = wb['Sheet1']  # сохранение одного листа таблицы
print(sheet.title)
print(sheet['A1'].value)
b2 = sheet['B2']
print(f'Line: {b2.row}, column: {b2.column}, value: {b2.value}')

high_row = sheet.max_row  # сохраняю номер последней строчки на листе
high_column = sheet.max_column  # сохраняю номер последней ячейки на листе
for i in range(1, high_row + 1):
    for j in range(1, high_column + 1):
        print(sheet.cell(row=i, column=j).value, end=" | ")
    print()
